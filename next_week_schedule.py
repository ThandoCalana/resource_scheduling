import os
import requests
import pytz
from datetime import datetime, timezone, timedelta, time
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# --- Load environment variables ---
load_dotenv()

# ClickUp
CLICKUP_API_TOKEN = os.getenv("CLICKUP_TOKEN")
SPACE_IDS = [s.strip() for s in os.getenv("CLICKUP_SPACE_IDS", "").split(",") if s.strip()]
ASSIGNEES = [a.strip() for a in os.getenv("CLICKUP_ASSIGNEES", "").split(",") if a.strip()]
ASSIGNEES_WITH_UNASSIGNED = ASSIGNEES + ["Unassigned"]
CLICKUP_HEADERS = {"Authorization": CLICKUP_API_TOKEN}


# Outlook
OUTLOOK_USER_EMAILS = [e.strip() for e in os.getenv("OUTLOOK_USER_EMAIL", "").split(",")]
TENANT_ID = os.getenv("TENANT_ID")
CLIENT_ID = os.getenv("CLIENT_ID")
CLIENT_SECRET = os.getenv("CLIENT_SECRET")
LOCAL_TZ = pytz.timezone("Africa/Johannesburg")

# --- Shared helpers ---
def get_week_dates():
    today = datetime.now(LOCAL_TZ).date()
    monday = today - timedelta(days=today.weekday())
    return [monday + timedelta(days=i) for i in range(5)]

def generate_time_slots(start_hour=8, end_hour=18):
    slots = []
    current = datetime.combine(datetime.today(), time(start_hour,0))
    end = datetime.combine(datetime.today(), time(end_hour,0))
    while current < end:
        slots.append(current.time())
        current += timedelta(minutes=30)
    return slots

def email_to_name(email): 
    return " ".join(p.capitalize() for p in email.split("@")[0].split("."))

# -------------------- CLICKUP  --------------------
def get_folders(space_id):
    return requests.get(f"https://api.clickup.com/api/v2/space/{space_id}/folder", headers=CLICKUP_HEADERS).json().get("folders", [])

def get_lists_in_folder(folder_id):
    return requests.get(f"https://api.clickup.com/api/v2/folder/{folder_id}/list", headers=CLICKUP_HEADERS).json().get("lists", [])

def get_lists_directly_in_space(space_id):
    return requests.get(f"https://api.clickup.com/api/v2/space/{space_id}/list", headers=CLICKUP_HEADERS).json().get("lists", [])

def get_tasks(list_id):
    return requests.get(f"https://api.clickup.com/api/v2/list/{list_id}/task?subtasks=true", headers=CLICKUP_HEADERS).json().get("tasks", [])

def get_subtasks(task_id):
    return requests.get(f"https://api.clickup.com/api/v2/task/{task_id}/subtask", headers=CLICKUP_HEADERS).json().get("tasks", [])


def fetch_clickup_tasks():
    now = datetime.now(timezone.utc)
    weekdays = get_week_dates()
    next_weekdays = [d + timedelta(days=28) for d in weekdays]
    excluded_lists = {"product management"}
    task_dict = {a: [] for a in ASSIGNEES_WITH_UNASSIGNED}
    seen = {a: set() for a in ASSIGNEES_WITH_UNASSIGNED}

    def build_sheet_dates(due, allow_overdue):
        if not due: return next_weekdays
        days = [d for d in next_weekdays if d <= due]
        return days or ([next_weekdays[0]] if allow_overdue else [])

    def push(task_id, name, link, assignees, sheet_dates):
        for a in assignees:
            target = a if a in ASSIGNEES_WITH_UNASSIGNED else "Unassigned"
            if task_id not in seen[target]:
                task_dict[target].append({"name": name, "sheet_dates": sheet_dates, "id": task_id, "link": link})
                seen[target].add(task_id)

    def add_task(t, allowed, list_name, restrict=False, is_sub=False, parent_due=None):
        tid = t.get("id")
        status = (t.get("status", {}) or {}).get("status", "").upper()
        due = datetime.fromtimestamp(int(t["due_date"])/1000, tz=timezone.utc).date() if t.get("due_date") else parent_due
        assignees = [a.get("username", "") for a in t.get("assignees", [])] or ["Unassigned"]
        allow_overdue = not restrict

        if restrict and ("Unassigned" in assignees or (due and due < now.date())):
            for sub in get_subtasks(tid): add_task(sub, allowed, list_name, restrict, True, due)
            return

        if status in allowed:
            sheet_dates = build_sheet_dates(due, allow_overdue)
            if sheet_dates:
                name = f"(Subtask) {t.get('name','Untitled')}" if is_sub else f"[{list_name}] {t.get('name','Untitled')}"
                push(tid, name, t.get("url"), assignees, sheet_dates)

        for sub in get_subtasks(tid): add_task(sub, allowed, list_name, restrict, True, due)

    for space_id in SPACE_IDS:
        for folder in get_folders(space_id):
            for lst in get_lists_in_folder(folder["id"]):
                lname = lst.get("name","").lower()
                if lname in excluded_lists: continue
                for t in get_tasks(lst["id"]):
                    if lname == "freshdesk": add_task(t, {"IN PROGRESS","TO DO","REVIEW"}, lname, restrict=False)
                    else: add_task(t, {"IN PROGRESS","REVIEW"}, lname, restrict=True)
        for lst in get_lists_directly_in_space(space_id):
            lname = lst.get("name","").lower()
            if lname in excluded_lists: continue
            for t in get_tasks(lst["id"]):
                if lname == "freshdesk": add_task(t, {"IN PROGRESS","TO DO","REVIEW"}, lname, restrict=False)
                else: add_task(t, {"IN PROGRESS","REVIEW"}, lname, restrict=True)

    return task_dict

# -------------------- OUTLOOK --------------------
def get_outlook_events(user):
    token_url = f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token"
    token_data = {
        "client_id": CLIENT_ID,
        "client_secret": CLIENT_SECRET,
        "scope": "https://graph.microsoft.com/.default",
        "grant_type": "client_credentials",
    }
    access_token = requests.post(token_url, data=token_data).json()["access_token"]
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Prefer": 'outlook.timezone="Africa/Johannesburg"',
    }

    # get this week's weekdays
    weekdays = get_week_dates()

    # shift them by 7 days to target next week (14 for 2 weeks from now etc.)
    next_weekdays = [d + timedelta(days=28) for d in weekdays]

    url = (
        f"https://graph.microsoft.com/v1.0/users/{user}/calendarview"
        f"?startDateTime={next_weekdays[0]}T00:00:00Z"
        f"&endDateTime={next_weekdays[-1]}T23:59:59Z&$top=1000"
    )

    events = requests.get(url, headers=headers).json().get("value", [])
    formatted = []
    for ev in events:
        start_dt = datetime.fromisoformat(ev["start"]["dateTime"]).astimezone(LOCAL_TZ)
        end_dt = datetime.fromisoformat(ev["end"]["dateTime"]).astimezone(LOCAL_TZ)
        formatted.append(
            {
                "subject": ev.get("subject", "No subject"),
                "date": start_dt.date(),
                "start_time": start_dt.time(),
                "end_time": end_dt.time(),
            }
        )
    return formatted

# -------------------- WRITE TO LOCAL EXCEL --------------------
def write_combined_excel(filename="Next_Week_Team_Schedule.xlsx"):
    if os.path.exists(filename):
        wb = load_workbook(filename)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1 and not wb.active["A1"].value:
            wb.remove(wb.active)

    all_events = {u: get_outlook_events(u) for u in OUTLOOK_USER_EMAILS}
    task_dict = fetch_clickup_tasks()
    weekdays = get_week_dates()
    next_weekdays = [d + timedelta(days=28) for d in weekdays]
    time_slots = generate_time_slots()

    for day in next_weekdays:
        sheet_name = f"{day.strftime('%A')} {day.isoformat()}"
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            wb.remove(ws)
        ws = wb.create_sheet(title=sheet_name)

        # Calendar rows
        rows = [["Time"] + [email_to_name(u) for u in OUTLOOK_USER_EMAILS]]
        for slot in time_slots:
            row = [slot.strftime("%H:%M")]
            for u in OUTLOOK_USER_EMAILS:
                evs = [ev["subject"] for ev in all_events[u] if ev["date"]==day and ev["start_time"]<=slot<ev["end_time"]]
                row.append(", ".join(evs))
            rows.append(row)

        # Leave 2 blank rows
        rows.append([""]*(len(OUTLOOK_USER_EMAILS)+1))
        rows.append([""]*(len(OUTLOOK_USER_EMAILS)+1))

        # --- Load % Row ---
        load_row = ["Load %"]
        start_index, end_index = None, None
        for idx, slot in enumerate(time_slots):
            if slot == time(8,0): start_index = idx + 1
            if slot == time(17,0): end_index = idx + 1
        if start_index is None: start_index = 1
        if end_index is None: end_index = len(time_slots)

        total_slots = end_index - start_index
        for col_idx in range(1, len(OUTLOOK_USER_EMAILS)+1):
            col_letter = get_column_letter(col_idx+1)
            formula = f'=ROUND(COUNTIF({col_letter}{start_index+1}:{col_letter}{end_index},"<>")/{total_slots},4)'
            load_row.append(formula)
        rows.append(load_row)

        # Leave 2 blank rows
        rows.append([""]*(len(OUTLOOK_USER_EMAILS)+1))
        rows.append([""]*(len(OUTLOOK_USER_EMAILS)+1))

        # Tasks header
        rows.append(["Tasks per Assignee"] + ASSIGNEES_WITH_UNASSIGNED)
        max_tasks = max([len(task_dict[a]) for a in ASSIGNEES_WITH_UNASSIGNED]+[0])
        for i in range(max_tasks):
            row = [""]
            for a in ASSIGNEES_WITH_UNASSIGNED:
                if i < len(task_dict[a]):
                    t = task_dict[a][i]
                    if t.get("link"):
                        row.append(f'=HYPERLINK("{t["link"]}", "{t["name"]}")')
                    else:
                        row.append(t["name"])
                else:
                    row.append("")
            rows.append(row)

        # Write to Excel
        for r_idx, row in enumerate(rows, 1):
            for c_idx, val in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=val)

        # Auto-fit columns
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    wb.save(filename)
    print(f"Local Excel written to {filename}")

# -------------------- MAIN --------------------
if __name__ == "__main__":
    print(datetime.now())
    write_combined_excel()
    print(datetime.now())
