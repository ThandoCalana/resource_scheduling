from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

SOURCE_FILE = "Three_Month_Team_Schedule.xlsx"
OUTPUT_FILE = "Transformed_Load_Output.xlsx"

wb_src = load_workbook(SOURCE_FILE, data_only=True)

day_sheets = [s for s in wb_src.sheetnames if s not in ["Names", "Daily Loads", "Monthly Averages"]]

first_sheet = wb_src[day_sheets[0]]
names = []
col = 2
while first_sheet.cell(row=1, column=col).value:
    names.append(first_sheet.cell(row=1, column=col).value)
    col += 1

daily_loads = {name: [] for name in names}
daily_dates = []

for sheet_name in day_sheets:
    ws = wb_src[sheet_name]
    date_part = sheet_name.split(" ")[-1]
    try:
        sheet_date = datetime.fromisoformat(date_part).date()
    except:
        continue
    daily_dates.append(sheet_date)
    for idx, name in enumerate(names):
        val = ws.cell(row=24, column=idx + 2).value
        if isinstance(val, (int, float)):
            daily_loads[name].append((sheet_date, val))

wb_out = Workbook()
ws_names = wb_out.active
ws_names.title = "Names"

for i, name in enumerate(names, 1):
    ws_names.cell(row=i, column=1, value=name)

ws_daily = wb_out.create_sheet("Daily Loads")
ws_daily.cell(row=1, column=1, value="Date")

for i, name in enumerate(names, 2):
    ws_daily.cell(row=1, column=i, value=name)

unique_dates = sorted({d for loads in daily_loads.values() for d, _ in loads})

for r, d in enumerate(unique_dates, 2):
    ws_daily.cell(row=r, column=1, value=d.isoformat())
    for c, name in enumerate(names, 2):
        match = [v for date, v in daily_loads[name] if date == d]
        if match:
            ws_daily.cell(row=r, column=c, value=match[0])

monthly = {}
for name, values in daily_loads.items():
    for d, v in values:
        key = d.strftime("%Y-%m")
        monthly.setdefault(name, {}).setdefault(key, []).append(v)

monthly_avg = {
    name: {m: sum(vals) / len(vals) for m, vals in months.items()}
    for name, months in monthly.items()
}

all_months = sorted({m for months in monthly_avg.values() for m in months})

ws_monthly = wb_out.create_sheet("Monthly Averages")

ws_monthly.cell(row=1, column=1, value="Name")
for c, m in enumerate(all_months, 2):
    ws_monthly.cell(row=1, column=c, value=datetime.strptime(m, "%Y-%m").strftime("%B %Y"))

for r, name in enumerate(names, 2):
    ws_monthly.cell(row=r, column=1, value=name)
    for c, m in enumerate(all_months, 2):
        if m in monthly_avg.get(name, {}):
            ws_monthly.cell(row=r, column=c, value=monthly_avg[name][m])

wb_out.save(OUTPUT_FILE)
