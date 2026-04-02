import os
import requests
import pytz
import pandas as pd
from datetime import datetime, timezone, timedelta, time
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

# -------------------- ENV SETUP --------------------
load_dotenv()

CLICKUP_API_TOKEN = os.environ["CLICKUP_TOKEN"]
SPACE_IDS = [s.strip() for s in os.environ["CLICKUP_SPACE_IDS"].split(",") if s.strip()]
ASSIGNEES = [a.strip() for a in os.environ["CLICKUP_ASSIGNEES"].split(",") if a.strip()]
ASSIGNEES_WITH_UNASSIGNED = ASSIGNEES + ["Unassigned"]

OUTLOOK_USER_EMAIL = [e.strip() for e in os.environ["OUTLOOK_USER_EMAIL"].split(",")]
TENANT_ID = os.environ["TENANT_ID"]
CLIENT_ID = os.environ["CLIENT_ID"]
CLIENT_SECRET = os.environ["CLIENT_SECRET"]

CLICKUP_HEADERS = {"Authorization": CLICKUP_API_TOKEN}

# Output paths
OUTPUT_DIR = "./data"
OUTPUT_CSV = os.path.join(OUTPUT_DIR, "calendar_flat.csv")

LOCAL_TZ = pytz.timezone("Africa/Johannesburg")

# -------------------- HELPERS --------------------
def get_week_dates():
    today = datetime.now(LOCAL_TZ).date()
    first_day = today.replace(day=1)
    return [first_day + timedelta(days=i) for i in range(92)]

def generate_time_slots(start_hour=8, end_hour=18):
    slots = []
    current = datetime.combine(datetime.now(LOCAL_TZ), time(start_hour, 0))
    end = datetime.combine(datetime.now(LOCAL_TZ), time(end_hour, 0))
    while current < end:
        slots.append(current.time())
        current += timedelta(minutes=30)
    return slots

def email_to_name(email):
    return " ".join(p.capitalize() for p in email.split("@")[0].split("."))

# -------------------- CLICKUP --------------------
def get_folders(space_id):
    return requests.get(f"https://api.clickup.com/api/v2/space/{space_id}/folder", headers=CLICKUP_HEADERS).json().get("folders", [])

def get_lists_in_folder(folder_id):
    return requests.get(f"https://api.clickup.com/api/v2/folder/{folder_id}/list", headers=CLICKUP_HEADERS).json().get("lists", [])

def get_lists_directly_in_space(space_id):
    return requests.get(f"https://api.clickup.com/api/v2/space/{space_id}/list", headers=CLICKUP_HEADERS).json().get("lists", [])

def get_tasks(list_id):
    return requests.get(f"https://api.clickup.com/api/v2/list/{list_id}/task?subtasks=true", headers=CLICKUP_HEADERS).json().get("tasks", [])

def get_subtasks(task_id):
    return requests.get(f"https://api.clickup.com/api/v2/task/{task_id}/subtask", headers=CLICKUP_HEADERS).json().get("tasks", [])

def fetch_clickup_tasks():
    now = datetime.now(timezone.utc)
    weekdays = get_week_dates()
    task_dict = {a: [] for a in ASSIGNEES_WITH_UNASSIGNED}
    seen = {a: set() for a in ASSIGNEES_WITH_UNASSIGNED}

    def build_sheet_dates(due, allow_overdue):
        if not due:
            return weekdays
        days = [d for d in weekdays if d <= due]
        return days or ([weekdays[0]] if allow_overdue else [])

    def push(task_id, name, link, assignees, sheet_dates):
        for a in assignees:
            target = a if a in ASSIGNEES_WITH_UNASSIGNED else "Unassigned"
            if task_id not in seen[target]:
                task_dict[target].append({
                    "name": name,
                    "sheet_dates": sheet_dates,
                    "id": task_id,
                    "link": link
                })
                seen[target].add(task_id)

    def add_task(t, allowed, list_name, restrict=False, is_sub=False, parent_due=None):
        tid = t.get("id")
        status = (t.get("status", {}) or {}).get("status", "").upper()
        due = datetime.fromtimestamp(int(t["due_date"]) / 1000, tz=timezone.utc).date() if t.get("due_date") else parent_due
        assignees = [a.get("username", "") for a in t.get("assignees", [])] or ["Unassigned"]

        if restrict and ("Unassigned" in assignees or (due and due < now.date())):
            for sub in get_subtasks(tid):
                add_task(sub, allowed, list_name, restrict, True, due)
            return

        if status in allowed:
            sheet_dates = build_sheet_dates(due, not restrict)
            if sheet_dates:
                name = f"(Subtask) {t.get('name','Untitled')}" if is_sub else f"[{list_name}] {t.get('name','Untitled')}"
                push(tid, name, t.get("url"), assignees, sheet_dates)

        for sub in get_subtasks(tid):
            add_task(sub, allowed, list_name, restrict, True, due)

    for space_id in SPACE_IDS:
        for folder in get_folders(space_id):
            for lst in get_lists_in_folder(folder["id"]):
                lname = lst.get("name", "").lower()
                for t in get_tasks(lst["id"]):
                    if lname == "freshdesk":
                        add_task(t, {"IN PROGRESS", "TO DO", "REVIEW"}, lname)
                    else:
                        add_task(t, {"IN PROGRESS", "REVIEW"}, lname, restrict=True)

        for lst in get_lists_directly_in_space(space_id):
            lname = lst.get("name", "").lower()
            for t in get_tasks(lst["id"]):
                if lname == "freshdesk":
                    add_task(t, {"IN PROGRESS", "TO DO", "REVIEW"}, lname)
                else:
                    add_task(t, {"IN PROGRESS", "REVIEW"}, lname, restrict=True)

    return task_dict

# -------------------- OUTLOOK --------------------
def get_outlook_events(user):
    token_url = f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token"
    token_data = {
        "client_id": CLIENT_ID,
        "client_secret": CLIENT_SECRET,
        "scope": "https://graph.microsoft.com/.default",
        "grant_type": "client_credentials",
    }

    access_token = requests.post(token_url, data=token_data).json()["access_token"]

    headers = {
        "Authorization": f"Bearer {access_token}",
        "Prefer": 'outlook.timezone="Africa/Johannesburg"',
    }

    weekdays = get_week_dates()

    url = (
        f"https://graph.microsoft.com/v1.0/users/{user}/calendarview"
        f"?startDateTime={weekdays[0]}T00:00:00Z"
        f"&endDateTime={weekdays[-1]}T23:59:59Z&$top=1000"
    )

    events = requests.get(url, headers=headers).json().get("value", [])

    formatted = []
    for ev in events:
        start_dt = datetime.fromisoformat(ev["start"]["dateTime"]).astimezone(LOCAL_TZ)
        end_dt = datetime.fromisoformat(ev["end"]["dateTime"]).astimezone(LOCAL_TZ)

        # filter useless events
        if (end_dt - start_dt).total_seconds() <= 0:
            continue

        formatted.append({
            "subject": ev.get("subject", "No subject"),
            "date": start_dt.date(),
            "start_time": start_dt.time(),
            "end_time": end_dt.time(),
        })

    return formatted

# -------------------- MAIN --------------------
def run_extraction():
    print(f"Started → {datetime.now(LOCAL_TZ)}")

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)

    all_events = {u: get_outlook_events(u) for u in OUTLOOK_USER_EMAIL}
    task_dict = fetch_clickup_tasks()

    weekdays = get_week_dates()
    time_slots = generate_time_slots()

    flat_rows = []

    for day in weekdays:
        sheet_name = f"{day.strftime('%A')} {day.isoformat()}"
        ws = wb.create_sheet(title=sheet_name)

        header = ["Time"] + [email_to_name(u) for u in OUTLOOK_USER_EMAIL]
        ws.append(header)

        for slot in time_slots:
            row = [slot.strftime("%H:%M")]

            for u in OUTLOOK_USER_EMAIL:
                user_name = email_to_name(u)

                evs = [
                    ev["subject"]
                    for ev in all_events[u]
                    if ev["date"] == day and ev["start_time"] <= slot < ev["end_time"]
                ]

                subject = ", ".join(evs)
                is_busy = 1 if subject else 0

                row.append(subject)

                flat_rows.append({
                    "date": day,
                    "time": slot.strftime("%H:%M"),
                    "user": user_name,
                    "subject": subject,
                    "is_busy": is_busy
                })

            ws.append(row)

    df = pd.DataFrame(flat_rows)
    df.to_csv(OUTPUT_CSV, index=False)

    print(f"Finished → {datetime.now(LOCAL_TZ)}")


if __name__ == "__main__":
    run_extraction()
