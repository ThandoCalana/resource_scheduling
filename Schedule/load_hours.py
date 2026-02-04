import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

# ----------------- User Settings -----------------
SOURCE_FILE = "Three_Month_Team_Schedule.xlsx"
OUTPUT_FILE = "Aggregated_Hours.xlsx"

REPORT_START = datetime(2026, 1, 1).date()
REPORT_END   = datetime(2026, 3, 31).date()
# -------------------------------------------------

# Load workbook
wb = load_workbook(SOURCE_FILE, data_only=True)
day_sheets = [s for s in wb.sheetnames if s not in ["Names", "Daily Loads", "Monthly Aggregation"]]

# --- Extract team members from first day sheet ---
first_sheet = wb[day_sheets[0]]
team_members = []
col = 2
while first_sheet.cell(row=1, column=col).value:
    team_members.append(first_sheet.cell(row=1, column=col).value)
    col += 1

team_members_df = pd.DataFrame({"Team Member": team_members})

# ---------------- Atomic fact table ----------------
facts = []

for sheet_name in day_sheets:
    try:
        sheet_date = datetime.fromisoformat(sheet_name.split(" ")[-1]).date()
    except:
        continue

    if not (REPORT_START <= sheet_date <= REPORT_END):
        continue

    if sheet_date.weekday() >= 5:
        continue  # exclude weekends

    df = pd.read_excel(SOURCE_FILE, sheet_name=sheet_name, header=None)
    headers = df.iloc[0, 1:].tolist()

    for row_idx in range(1, 21):  # rows 2–21 = 30-min slots
        for member in team_members:
            if member not in headers:
                continue

            col_idx = headers.index(member) + 1
            task = df.iat[row_idx, col_idx]

            if pd.notna(task):
                facts.append({
                    "Date": pd.to_datetime(sheet_date),
                    "Team Member": member,
                    "Task": str(task).strip(),
                    "Hours": 0.5
                })

fact_df = pd.DataFrame(facts)

# ---------------- Sheet 2: Daily Loads ----------------
daily = (
    fact_df
    .groupby(["Date", "Team Member"])["Hours"]
    .sum()
    .reset_index()
)

daily_pivot = (
    daily
    .pivot(index="Date", columns="Team Member", values="Hours")
    .fillna(0)
    .reindex(columns=team_members, fill_value=0)
    .sort_index()
)

# ---------------- Sheet 3: Monthly Aggregation ----------------
fact_df["MonthDate"] = fact_df["Date"].dt.to_period("M").dt.to_timestamp()

monthly = (
    fact_df
    .groupby(["Team Member", "MonthDate"])["Hours"]
    .sum()
    .reset_index()
)

monthly_pivot = (
    monthly
    .pivot(index="Team Member", columns="MonthDate", values="Hours")
    .fillna(0)
    .reindex(index=team_members, fill_value=0)
    .sort_index(axis=1)
)

monthly_pivot.columns = [d.strftime("%B %Y") for d in monthly_pivot.columns]

# -------- Per-member task monthly aggregation sheets --------
member_task_monthly = (
    fact_df
    .groupby(["Team Member", "Task", "MonthDate"])["Hours"]
    .sum()
    .reset_index()
)

# ---------------- Write to Excel ----------------
with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    team_members_df.to_excel(writer, sheet_name="Names", index=False)
    daily_pivot.to_excel(writer, sheet_name="Daily Loads")
    monthly_pivot.to_excel(writer, sheet_name="Monthly Aggregation")

    for member in team_members:
        member_df = member_task_monthly[
            member_task_monthly["Team Member"] == member
        ]

        if member_df.empty:
            continue

        pivot = (
            member_df
            .pivot(index="Task", columns="MonthDate", values="Hours")
            .fillna(0)
            .sort_index(axis=1)  # 🔑 true chronological ordering
        )

        # Format month headers AFTER ordering is fixed
        pivot.columns = [d.strftime("%B %Y") for d in pivot.columns]

        pivot.to_excel(writer, sheet_name=member[:31])


print(f"Output written to {OUTPUT_FILE}")
